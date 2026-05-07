"""
c3_twin_loader — structured reader for the C3 Digital Twin extraction XLSX.

Typical use:

    from c3_twin_loader import C3Twin
    tw = C3Twin()                                 # uses default XLSX in CWD

    tw.species("C3")                              # all rows for C3
    tw.species_value("C3", paper="ZEWDE_2016_AP_core")          # -> 5.4
    tw.initial_conditions(condition="healthy_baseline")          # {name: uM}
    tw.kinetic("R003a")                           # KineticParam dataclass
    tw.validation("C3a", disease_state="AMD_Y402H")              # [DataPoint,...]
    tw.bounds("k_C3b_FB_bind")                    # (lower, upper, unit)
    tw.drug("Eculizumab")                         # {param_name: DrugParam}
    tw.perturbation("PNH_conv_decay_0.25x")       # PerturbationRule
    tw.apply_perturbation(name, params)           # returns modified-params dict
    tw.steady_state("C3", model_variant="minimal_model")

No pandas dependency — pure-Python + openpyxl. All dictionaries are plain dicts
so the loader can be dropped into a JAX / NumPy / SciPy ODE pipeline without
extra dependencies.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = Path(__file__).parent / "C3_Digital_Twin_Literature_Extraction_Template.xlsx"


# =============================================================================
# Unit normalisation helpers
# =============================================================================

# All concentrations are normalised to micromolar (uM) internally.
_CONC_FACTORS_TO_UM = {
    "m": 1e6, "M": 1e6,
    "mm": 1e3, "mM": 1e3,
    "um": 1.0, "uM": 1.0, "μM": 1.0,
    "nm": 1e-3, "nM": 1e-3,
    "pm": 1e-6, "pM": 1e-6,
}

# All rate constants are normalised to 1/min and 1/(uM*min) internally.
_TIME_TO_MIN = {
    "s": 1 / 60, "sec": 1 / 60, "second": 1 / 60,
    "min": 1.0, "minute": 1.0,
    "h": 60.0, "hr": 60.0, "hour": 60.0,
    "day": 60.0 * 24,
}


def conc_to_uM(value: float | None, unit: str | None) -> float | None:
    """Convert a concentration to uM. Returns None if either input is None."""
    if value is None or unit is None:
        return None
    factor = _CONC_FACTORS_TO_UM.get(unit.strip())
    return value * factor if factor is not None else None


def time_to_min(value: float | None, unit: str | None) -> float | None:
    if value is None or unit is None:
        return None
    factor = _TIME_TO_MIN.get(unit.strip().lower())
    return value * factor if factor is not None else None


# =============================================================================
# Record dataclasses
# =============================================================================

@dataclass(slots=True)
class Species:
    paper: str
    name: str
    compartment: str | None
    baseline: float | None
    unit: str | None
    mw_kDa: float | None = None
    half_life: float | None = None
    half_life_unit: str | None = None
    condition: str | None = None
    source: str | None = None
    notes: str | None = None

    @property
    def baseline_uM(self) -> float | None:
        return conc_to_uM(self.baseline, self.unit)


@dataclass(slots=True)
class KineticParam:
    paper: str
    reaction_id: str
    reactants: str | None
    products: str | None
    parameter_name: str | None
    parameter_type: str | None
    value: float | None
    unit: str | None
    context: str | None = None
    source: str | None = None
    notes: str | None = None


@dataclass(slots=True)
class ValidationPoint:
    paper: str
    observable: str
    compartment: str | None
    t_or_dose: float | None
    t_or_dose_unit: str | None
    value: float | None
    value_unit: str | None
    condition: str | None
    disease_state: str | None = None
    intervention: str | None = None
    intervention_dose: str | None = None
    model_variant: str | None = None
    source: str | None = None
    notes: str | None = None

    @property
    def t_min(self) -> float | None:
        return time_to_min(self.t_or_dose, self.t_or_dose_unit)


@dataclass(slots=True)
class PerturbationRule:
    paper: str
    name: str
    target: str
    perturbation_type: str
    numeric_value: str
    numeric_unit: str | None = None
    disease_state: str | None = None
    intervention: str | None = None
    biological_meaning: str | None = None
    notes: str | None = None

    @property
    def factor(self) -> float | None:
        """Best-effort interpretation of numeric_value as a multiplicative factor."""
        v = self.numeric_value
        if v is None:
            return None
        try:
            x = float(v)
        except (TypeError, ValueError):
            return None
        t = (self.perturbation_type or "").lower()
        if t in ("multiply", "upregulation", "downregulation", "set"):
            # "downregulation 10-fold" in this template means x=10, factor=1/10
            if t == "downregulation":
                return 1 / x if x else None
            return x
        return x


@dataclass(slots=True)
class Bound:
    item_name: str
    bound_type: str
    lower: float | None
    upper: float | None
    unit: str | None
    paper: str
    context: str | None = None


@dataclass(slots=True)
class SteadyState:
    paper: str
    species: str
    model_variant: str | None
    value: float | None
    unit: str | None
    condition: str | None = None
    notes: str | None = None


@dataclass(slots=True)
class DrugParam:
    paper: str
    modality: str
    parameter_name: str
    value: float | None
    unit: str | None
    compartment: str | None = None
    context: str | None = None
    notes: str | None = None


# =============================================================================
# Main loader
# =============================================================================

class C3Twin:
    def __init__(self, xlsx_path: str | Path = DEFAULT_XLSX):
        self.path = Path(xlsx_path)
        if not self.path.exists():
            raise FileNotFoundError(self.path)
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        self._load(wb)
        wb.close()

    # ---- low-level sheet ingestion ----
    @staticmethod
    def _iter_sheet(ws) -> list[dict[str, Any]]:
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            rec = {h: row[i] for i, h in enumerate(headers) if h}
            # Skip placeholder rows: paper_id present but no payload
            paper = rec.get("paper_id")
            if not paper:
                continue
            payload_keys = [k for k in rec
                            if k not in ("paper_id", "source_location", "source_url", "notes")]
            if not any(rec.get(k) not in (None, "") for k in payload_keys):
                continue
            out.append(rec)
        return out

    def _load(self, wb: openpyxl.Workbook) -> None:
        sp_rec = self._iter_sheet(wb["Template_species_baseline"])
        self.species_records: list[Species] = [
            Species(
                paper=r["paper_id"], name=r["species_name"],
                compartment=r.get("compartment"),
                baseline=r.get("baseline_value"), unit=r.get("unit"),
                mw_kDa=r.get("molecular_weight_kDa"),
                half_life=r.get("half_life"),
                half_life_unit=r.get("half_life_unit"),
                condition=r.get("condition_label"),
                source=r.get("source_location"), notes=r.get("notes"),
            )
            for r in sp_rec
        ]

        kp_rec = self._iter_sheet(wb["Template_kinetic_parameters"])
        self.kinetic_records: list[KineticParam] = [
            KineticParam(
                paper=r["paper_id"], reaction_id=r["reaction_id"],
                reactants=r.get("reactants"), products=r.get("products"),
                parameter_name=r.get("parameter_name"),
                parameter_type=r.get("parameter_type"),
                value=r.get("value"), unit=r.get("unit"),
                context=r.get("context"),
                source=r.get("source_location"), notes=r.get("notes"),
            )
            for r in kp_rec
        ]

        vt_rec = self._iter_sheet(wb["Template_validation_timeseries"])
        self.validation_records: list[ValidationPoint] = [
            ValidationPoint(
                paper=r["paper_id"], observable=r["observable_name"],
                compartment=r.get("compartment_or_surface"),
                t_or_dose=r.get("time_or_dose"),
                t_or_dose_unit=r.get("time_or_dose_unit"),
                value=r.get("value"), value_unit=r.get("value_unit"),
                condition=r.get("condition"),
                disease_state=r.get("disease_state"),
                intervention=r.get("intervention_name"),
                intervention_dose=r.get("intervention_dose"),
                model_variant=r.get("model_variant"),
                source=r.get("source_location"), notes=r.get("notes"),
            )
            for r in vt_rec
        ]

        pr_rec = self._iter_sheet(wb["Template_perturbation_rules"])
        self.perturbation_records: list[PerturbationRule] = [
            PerturbationRule(
                paper=r["paper_id"], name=r["perturbation_name"],
                target=r.get("target"),
                perturbation_type=r.get("perturbation_type"),
                numeric_value=r.get("numeric_value"),
                numeric_unit=r.get("numeric_unit"),
                disease_state=r.get("disease_state"),
                intervention=r.get("intervention_name"),
                biological_meaning=r.get("biological_meaning"),
                notes=r.get("notes"),
            )
            for r in pr_rec
        ]

        pb_rec = self._iter_sheet(wb["Template_parameter_bounds"])
        self.bound_records: list[Bound] = [
            Bound(
                item_name=r["item_name"],
                bound_type=r.get("bound_type"),
                lower=r.get("lower_bound"), upper=r.get("upper_bound"),
                unit=r.get("unit"),
                paper=r["paper_id"],
                context=r.get("analysis_context"),
            )
            for r in pb_rec
        ]

        ss_rec = self._iter_sheet(wb["Template_steady_state_targets"])
        self.steady_records: list[SteadyState] = [
            SteadyState(
                paper=r["paper_id"], species=r["species_name"],
                model_variant=r.get("model_variant"),
                value=r.get("steady_state_value"), unit=r.get("unit"),
                condition=r.get("condition"),
                notes=r.get("notes"),
            )
            for r in ss_rec
        ]

        dp_rec = self._iter_sheet(wb["Template_drug_pk_parameters"])
        self.drug_records: list[DrugParam] = [
            DrugParam(
                paper=r["paper_id"], modality=r["drug_modality"],
                parameter_name=r.get("parameter_name"),
                value=r.get("value"), unit=r.get("unit"),
                compartment=r.get("compartment"),
                context=r.get("context"), notes=r.get("notes"),
            )
            for r in dp_rec
        ]

    # =========================================================================
    # Query API
    # =========================================================================
    def species(self, name: str, paper: str | None = None) -> list[Species]:
        return [s for s in self.species_records
                if s.name == name and (paper is None or s.paper == paper)]

    def species_value(self, name: str, paper: str | None = None,
                      prefer: list[str] | None = None) -> float | None:
        """Return first non-null baseline value (uM) matching the filters.

        If `prefer` is a list of paper_ids, look through them in order; otherwise
        take the first matching species row with a non-null baseline.
        """
        cands = self.species(name, paper=paper)
        if prefer:
            order = {p: i for i, p in enumerate(prefer)}
            cands.sort(key=lambda s: order.get(s.paper, 1_000))
        for s in cands:
            v = s.baseline_uM
            if v is not None:
                return v
        return None

    def initial_conditions(self, condition: str = "healthy_baseline",
                           prefer: list[str] | None = None) -> dict[str, float]:
        """Return {species_name: baseline_uM} for a given condition label."""
        prefer = prefer or ["ZEWDE_2016_AP_core", "BANSAL_2022_QSP", "CARUSO_2020_hemolysis"]
        result: dict[str, float] = {}
        for s in self.species_records:
            if s.condition != condition:
                continue
            v = s.baseline_uM
            if v is None:
                continue
            # keep the value from the highest-preference paper
            if s.name in result:
                existing_paper = next(
                    (r.paper for r in self.species_records
                     if r.name == s.name and r.condition == condition
                     and r.baseline_uM == result[s.name]), None)
                if existing_paper in prefer and s.paper in prefer:
                    if prefer.index(existing_paper) <= prefer.index(s.paper):
                        continue
            result[s.name] = v
        return result

    def kinetic(self, reaction_id: str) -> KineticParam | None:
        for k in self.kinetic_records:
            if k.reaction_id == reaction_id:
                return k
        return None

    def kinetics_by_context(self, context: str) -> list[KineticParam]:
        return [k for k in self.kinetic_records if k.context == context]

    def kinetic_value(self, reaction_id: str) -> float | None:
        k = self.kinetic(reaction_id)
        return k.value if k else None

    def validation(self, observable: str, *,
                   disease_state: str | None = None,
                   condition: str | None = None,
                   intervention: str | None = None) -> list[ValidationPoint]:
        out = [v for v in self.validation_records if v.observable == observable]
        if disease_state is not None:
            out = [v for v in out if v.disease_state == disease_state]
        if condition is not None:
            out = [v for v in out if v.condition == condition]
        if intervention is not None:
            out = [v for v in out if v.intervention == intervention]
        out.sort(key=lambda v: (v.t_min if v.t_min is not None
                                else (v.t_or_dose if v.t_or_dose is not None else 0)))
        return out

    def validation_curve(self, observable: str, **filters) -> tuple[list[float], list[float]]:
        """Return (times_min_or_doses, values) as parallel lists."""
        pts = self.validation(observable, **filters)
        ts, vs = [], []
        for p in pts:
            t = p.t_min if p.t_or_dose_unit in _TIME_TO_MIN else p.t_or_dose
            if t is None or p.value is None:
                continue
            ts.append(t)
            vs.append(p.value)
        return ts, vs

    def bounds(self, item_name: str) -> Bound | None:
        # prefer rows with at least one numeric bound over placeholder rows
        for b in self.bound_records:
            if b.item_name == item_name and (b.lower is not None or b.upper is not None):
                return b
        for b in self.bound_records:
            if b.item_name == item_name:
                return b
        return None

    def steady_state(self, species: str, model_variant: str | None = None,
                     condition: str | None = None) -> SteadyState | None:
        for s in self.steady_records:
            if s.species != species:
                continue
            if model_variant is not None and s.model_variant != model_variant:
                continue
            if condition is not None and s.condition != condition:
                continue
            return s
        return None

    def perturbation(self, name: str) -> PerturbationRule | None:
        for p in self.perturbation_records:
            if p.name == name:
                return p
        return None

    def apply_perturbation(self, name: str, params: dict[str, float]) -> dict[str, float]:
        """Return a new params dict with the perturbation applied to `target`.

        Parameters
        ----------
        name : perturbation rule name (e.g. 'PNH_conv_decay_0.25x')
        params : base parameter dict {target: value}

        The target in a PerturbationRule may be a free-text phrase ("FH
        concentration and FH association rate"); when it maps to multiple keys
        in `params`, all matching keys get the factor applied.
        """
        p = self.perturbation(name)
        if p is None:
            raise KeyError(f"unknown perturbation {name!r}")
        factor = p.factor
        if factor is None:
            raise ValueError(f"perturbation {name!r} has non-numeric factor {p.numeric_value!r}")
        keys = [k for k in params if k in (p.target or "")] or [p.target]
        out = dict(params)
        for k in keys:
            if k in out and out[k] is not None:
                out[k] = out[k] * factor
        return out

    def drug(self, drug_label: str) -> dict[str, DrugParam]:
        """Return {parameter_name: DrugParam} for every row whose notes/context
        mentions the drug label (case-insensitive substring match)."""
        lab = drug_label.lower()
        out: dict[str, DrugParam] = {}
        for d in self.drug_records:
            haystack = " ".join(filter(None, [d.context or "", d.notes or "",
                                              d.parameter_name or ""])).lower()
            if lab in haystack:
                out[d.parameter_name] = d
        return out

    # ---- convenience ----
    def summary(self) -> dict[str, int]:
        return {
            "species_rows": len(self.species_records),
            "kinetic_rows": len(self.kinetic_records),
            "validation_rows": len(self.validation_records),
            "perturbation_rows": len(self.perturbation_records),
            "bound_rows": len(self.bound_records),
            "steady_rows": len(self.steady_records),
            "drug_rows": len(self.drug_records),
        }


# =============================================================================
# Smoke test
# =============================================================================
if __name__ == "__main__":
    tw = C3Twin()
    print("Loaded:", tw.summary())

    print("\n-- species('C3') -> all rows --")
    for s in tw.species("C3"):
        print(f"  {s.paper:30s} baseline={s.baseline} {s.unit}  mw={s.mw_kDa}")

    print("\n-- species_value('C3') with preference order --")
    print("  ", tw.species_value("C3"))

    print("\n-- initial_conditions('healthy_baseline') — first 8 items --")
    ic = tw.initial_conditions()
    for n, v in list(ic.items())[:8]:
        print(f"  {n:20s} {v} uM")
    print(f"  ... total {len(ic)} species")

    print("\n-- kinetic('R003a') — core C3 convertase catalysis --")
    k = tw.kinetic("R003a")
    print(" ", k)

    print("\n-- validation('C3a', disease_state='AMD_Y402H') --")
    for p in tw.validation("C3a", disease_state="AMD_Y402H"):
        print(f"  t={p.t_or_dose}{p.t_or_dose_unit}  value={p.value}{p.value_unit}  cond={p.condition}")

    print("\n-- bounds('k_C3b_FB_bind') --")
    print(" ", tw.bounds("k_C3b_FB_bind"))

    print("\n-- steady_state('C3', model_variant='properdin_model') --")
    print(" ", tw.steady_state("C3", model_variant="properdin_model"))

    print("\n-- drug('Eculizumab') — 4 key params --")
    for name, d in list(tw.drug("eculizumab").items())[:6]:
        print(f"  {name:20s} {d.value} {d.unit}")

    print("\n-- perturbation('PNH_conv_decay_0.25x') --")
    p = tw.perturbation("PNH_conv_decay_0.25x")
    print(" ", p)
    print("  factor =", p.factor)

    print("\n-- apply_perturbation on {k_conv_decay: 0.462} --")
    print(" ", tw.apply_perturbation("PNH_conv_decay_0.25x", {"k_conv_decay": 0.462}))
